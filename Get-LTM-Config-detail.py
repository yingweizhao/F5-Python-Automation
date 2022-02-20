# -*- coding: utf-8 -*-
import openpyxl
import bigsuds
def get_vs():
    vss=b.LocalLB.VirtualServer.get_list()
    return vss
def get_vip(vs):
    vip_info_list = b.LocalLB.VirtualServer.get_destination_v2([vs])
    vs_limite=b.LocalLB.VirtualServer.get_connection_limit([vs])
    vip = str(vip_info_list[0]['address']).split('/Common/')[1] + ':' + str(vip_info_list[0]['port'])
    return  vip,vs_limite
def get_protocol(vs):
    protocol = b.LocalLB.VirtualServer.get_protocol([vs])
    return protocol[0]
def get_irules(vs):
    irules = b.LocalLB.VirtualServer.get_rule([vs])
    irules_name=''
    irules_content=''
    for r in irules:
        if not len(r) == 0:
            irules_name = r[0]['rule_name']
            irules_content = b.LocalLB.Rule.query_rule([irules_name])
            irules_content= irules_content[0]['rule_definition']
    return irules_name,irules_content
def get_persistence(vs):
    persistence_info = b.LocalLB.VirtualServer.get_persistence_profile([vs])
    persistence=''
    if not len(persistence_info[0])==0:
       persistence=persistence_info[0][0]['profile_name']
    return persistence
def get_profile(vs):
    profile_info_list = b.LocalLB.VirtualServer.get_profile([vs])
    profile= ''
    for p in profile_info_list[0]:
        profile= profile + p['profile_name'] + '\n'
    return profile
def get_pool_name_member(vs):
    pool_name_info = b.LocalLB.VirtualServer.get_default_pool_name([vs])
    if not pool_name_info[0] == '':
        pool_name= pool_name_info[0]
    sub_m = ''
    if not pool_name_info[0] == '':
        pool_members = b.LocalLB.Pool.get_member([pool_name])
        for m in pool_members[0]:
            if not len(pool_members[0]) == 0:
                sub_m = sub_m + str(m['address']) + ':' + str(m['port']) +'\n'
        return pool_name,sub_m
    else:
         return None,None
def get_pool_from_irules(irules_info):
  pool_list=[]
  for s in str(irules_info).split("\n"):
    if s.find('pool')>=1:
        s=s.lstrip(' ')
        pool_list.append(str(s).split(' ')[1])
  return pool_list
def handle_pool_list(pool_list):
    pool_str=''
    if len(get_pool_from_irules(irules_info[1]))>=1:
        pools=get_pool_from_irules(irules_info[1])
        for pool in pools:
            pool='/Common/'+pool
            pool_str=pool_str+pool+'\n'
        #print pool_str
        return pool_str
    else:
        pool_str=''
        return pool_str
def get_pool_monitor(pool_name):
    pool_monitor=''
    if pool_name !='':
       pool_monitor_info_list = b.LocalLB.Pool.get_monitor_instance([pool_name])
       for pool_monitor_info in pool_monitor_info_list:
           if len(pool_monitor_info) !=0:
             pool_monitor = pool_monitor_info[0]['instance']['template_name']
       return pool_monitor
    else:
        return ''
def handle_irules_pool_monitor(irules_pool_list):
    if irules_pool_list !='':
       pool_monitor_list=''
       for poolName in str(irules_pool_list).split('\n'):
           pool_moitor=get_pool_monitor(poolName)
           pool_monitor_list=pool_monitor_list+pool_moitor+'\n'
       #print pool_monitor_list
       return pool_monitor_list


def set_sheet():
    sheet = wb.active
    sheet['A1'].value='VS NAME'
    sheet['B1'].value = 'VIP'
    sheet['C1'].value = 'VS limite'
    sheet['D1'].value = 'Protocol'
    sheet['E1'].value = 'Persistence'
    sheet['F1'].value = 'Profile'
    sheet['G1'].value = 'iRules Name'
    sheet['H1'].value = 'iRules Content'
    sheet['I1'].value = 'iRules Pool'
    sheet['J1'].value = 'iRules Monitor'
    sheet['K1'].value = 'Pool Name'
    sheet['L1'].value = 'Pool Monitor'
    sheet['M1'].value = 'Pool Member'
    return  sheet
def write_excel_content(vs,vip,vs_limite,protocol,persistence,profile,\
        irules_name,irules_content,pool_name,pool_member,line,irules_pool_list,irules_pool_monitor,pool_monitor):
    sheet=set_sheet()
    for row in range(line, line + 1):
        sheet.title = u'F5_LTM配置表'
        sheet.cell(row=row, column=1).value =str(vs).split('/Common/')[1]
        sheet.cell(row=row, column=2).value = vip
        sheet.cell(row=row, column=3).value = vs_limite
        sheet.cell(row=row, column=4).value = protocol
        sheet.cell(row=row, column=5).value = persistence
        sheet.cell(row=row, column=6).value = profile
        sheet.cell(row=row, column=7).value = irules_name
        sheet.cell(row=row, column=8).value = irules_content
        sheet.cell(row=row, column=9).value = irules_pool_list
        sheet.cell(row=row, column=10).value = irules_pool_monitor
        sheet.cell(row=row, column=11).value = pool_name
        sheet.cell(row=row, column=12).value = pool_monitor
        sheet.cell(row=row, column=13).value = pool_member

if __name__ == '__main__':
 global b,wb,save_file_name
 b=bigsuds.BIGIP(hostname='10.1.1.211',username='admin',password='admin')
 print "*************************"
 print "starting get F5 BIGIP LTM Config"
 wb = openpyxl.Workbook()
 save_file_name=u'F5_BIGIP_LTM10.0.4.15.xlsx'
 vss=get_vs()
 for i in range(len(vss)):
    vs=vss[i]
    #print i,vs
    vip,vs_limite=get_vip(vs)
    protocol=get_protocol(vs)
    persistence=get_persistence(vs)
    profile=get_profile(vs)
    irules_info=get_irules(vs)
    irules_name=irules_info[0]
    irules_content=irules_info[1]
    irules_pool_list=handle_pool_list(get_pool_from_irules(irules_info))
    irules_pool_monitor=handle_irules_pool_monitor(irules_pool_list)
    pool_info=get_pool_name_member(vs)
    pool_name=pool_info[0]
    pool_monitor = get_pool_monitor(pool_name)
    pool_member=pool_info[1]
    write_excel_content(vs,vip,vs_limite,protocol,persistence,profile,irules_name,irules_content,\
                        pool_name,pool_member,i+2,irules_pool_list,irules_pool_monitor,pool_monitor)
wb.save(save_file_name)
print "F5 BIGIP Config get finished!"



